"""
Import NBF_ORDER_At_A_Glance.xlsx into purchase_orders and parties tables.
Reads all "Order Details" sheets across all FY tabs.
Usage: python3 scripts/import_nbf_excel.py
Requires: MGMT and REF env vars (same as run_sql.py)
"""
import sys, os, json, subprocess, tempfile, datetime
from pathlib import Path

MGMT = os.environ["MGMT"]
REF  = os.environ.get("REF", "kjliulgpipqqwptinrrd")
XLSX = Path(__file__).parent.parent / "data" / "NBF_ORDER_At_A_Glance.xlsx"

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)
    import openpyxl

# ── helpers ──────────────────────────────────────────────────────────────────

def sq(v):
    """SQL single-quote escape."""
    if v is None:
        return "NULL"
    s = str(v).strip().replace("'", "''")
    return f"'{s}'"

def snum(v):
    if v is None:
        return "NULL"
    try:
        f = float(str(v).replace(",", ""))
        return str(f) if f != int(f) else str(int(f))
    except (ValueError, TypeError):
        return "NULL"

def sdate(v):
    if v is None:
        return "NULL"
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).strip()
    if not s:
        return "NULL"
    # dd-Mon-yy or dd-Mon-yyyy
    MON = {"jan":"01","feb":"02","mar":"03","apr":"04","may":"05","jun":"06",
           "jul":"07","aug":"08","sep":"09","oct":"10","nov":"11","dec":"12"}
    import re
    m = re.match(r'^(\d{1,2})[\/\-]([A-Za-z]{3})[\/\-](\d{2,4})$', s)
    if m:
        mo = MON.get(m.group(2).lower(), "01")
        yr = m.group(3) if len(m.group(3)) == 4 else f"20{m.group(3)}"
        return f"'{yr}-{mo}-{m.group(1).zfill(2)}'"
    m2 = re.match(r'^(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{2,4})$', s)
    if m2:
        yr = m2.group(3) if len(m2.group(3)) == 4 else f"20{m2.group(3)}"
        return f"'{yr}-{m2.group(2).zfill(2)}-{m2.group(1).zfill(2)}'"
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return f"'{s}'"
    return "NULL"

def fy_from_date(v):
    d = None
    if isinstance(v, (datetime.datetime, datetime.date)):
        d = v
    elif v:
        try:
            import re
            s = str(v).strip()
            m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
            if m:
                d = datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            pass
    if d is None:
        return None
    mo = d.month if isinstance(d, datetime.date) else d.strftime("%m")
    mo = int(mo) if isinstance(mo, str) else mo
    y  = d.year if isinstance(d, datetime.date) else int(d.strftime("%Y"))
    return f"{y}-{str(y+1)[2:]}" if mo >= 4 else f"{y-1}-{str(y)[2:]}"

def fy_from_sheet(sname):
    import re
    m = re.search(r'(\w{3,6})(\d{2})[-\s–]+(\w{3,6})(\d{2})', sname)
    if m:
        y1, y2 = m.group(2), m.group(4)
        return f"20{y1}-{y2}"
    m2 = re.search(r'(\d{4}).*?(\d{2,4})', sname)
    if m2:
        y1, y2 = m2.group(1), m2.group(2)
        y2 = y2 if len(y2) == 4 else f"20{y2}"
        return f"{y1}-{y2[2:]}"
    return "Unknown"

def norm_gst(v):
    if v is None:
        return 0
    try:
        f = float(str(v).replace(",","").replace("%",""))
        return round(f * 100, 2) if f <= 1 else round(f, 2)
    except (ValueError, TypeError):
        return 0

def norm_po_no(v):
    if not v:
        return None
    s = str(v).strip()
    # Normalise spaces around /
    import re
    s = re.sub(r'\s*/\s*', ' / ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ── parse Excel ──────────────────────────────────────────────────────────────

def parse_excel(path):
    wb = openpyxl.load_workbook(str(path), data_only=True)
    po_records  = []   # for purchase_orders
    vendors     = {}   # name → {category hints}

    for sname in wb.sheetnames:
        if "order details" not in sname.lower():
            continue

        ws = wb[sname]
        sheet_fy = fy_from_sheet(sname)
        rows = list(ws.iter_rows(min_row=1, values_only=True))

        # Detect header row (contains DATE OF INDENT or ITEM DESCRIPTION)
        header_row_idx = -1
        for i, row in enumerate(rows[:20]):
            vals = [str(v).upper() if v else "" for v in row]
            if any("INDENT" in v or "ITEM" in v for v in vals):
                header_row_idx = i
                break
        if header_row_idx < 0:
            continue

        # Find actual column header row (may be 1-2 below title)
        # It contains DATE OF INDENT / ITEM DESCRIPTION
        col_hdr_idx = header_row_idx
        for i in range(header_row_idx, min(header_row_idx + 4, len(rows))):
            row = rows[i]
            vals = [str(v).upper().strip() if v else "" for v in row]
            if any("DATE OF INDENT" in v or "ITEM DESCRIPTION" in v or "NAME OF ITEM" in v for v in vals):
                col_hdr_idx = i
                break

        hdr = [str(v).upper().strip() if v else "" for v in rows[col_hdr_idx]]

        # Detect layout variant
        # Old (22-25): UOM col comes before PACK SIZE
        # New (25-27): PACK SIZE col comes before UOM
        uom_idx   = next((i for i,h in enumerate(hdr) if "UOM" in h), -1)
        pack_idx  = next((i for i,h in enumerate(hdr) if "PACK" in h), -1)
        item_idx  = next((i for i,h in enumerate(hdr) if "ITEM" in h or "DESCRIPTION" in h), -1)
        date_idx  = next((i for i,h in enumerate(hdr) if "INDENT" in h or "DATE" in h), -1)
        qty_idx   = next((i for i,h in enumerate(hdr) if "ORDER" in h and "QTY" in h.replace(" ","")), -1)
        req_idx   = next((i for i,h in enumerate(hdr) if "REQUIRE" in h), -1)
        price_idx = next((i for i,h in enumerate(hdr) if "PRICE" in h and "NET" not in h), -1)
        gst_p_idx = next((i for i,h in enumerate(hdr) if "GST" in h and "%" in h), -1)
        gst_c_idx = next((i for i,h in enumerate(hdr) if "GST" in h and ("CRG" in h or "CHR" in h)), -1)
        net_idx   = next((i for i,h in enumerate(hdr) if "NET" in h), -1)
        po_date_i = next((i for i,h in enumerate(hdr) if "PO RAISED" in h or "PO CODE" in h), -1)
        po_no_idx = next((i for i,h in enumerate(hdr) if "PO DATE" in h or "PO CODE DATE" in h), -1)
        vendor_idx= next((i for i,h in enumerate(hdr) if "SUPPLIER" in h), -1)

        # Fallback fixed indices for known layouts
        if item_idx < 0:  item_idx  = 3
        if date_idx < 0:  date_idx  = 1
        if vendor_idx < 0: vendor_idx = 19

        # Old layout fixed fallbacks
        if uom_idx < 0 and pack_idx >= 0:
            # new layout: pack before qty before uom
            uom_idx = pack_idx + 2
        if qty_idx < 0:
            # ORDER QTY is typically right after PACK SIZE
            qty_idx = pack_idx + 1 if pack_idx >= 0 else 8
        if price_idx < 0: price_idx = 10
        if gst_p_idx < 0: gst_p_idx = 11
        if gst_c_idx < 0: gst_c_idx = 12
        if net_idx < 0:   net_idx   = 13
        if po_date_i < 0: po_date_i = 15
        if po_no_idx < 0: po_no_idx = 17

        data_start = col_hdr_idx + 2

        for row in rows[data_start:]:
            if not row or not any(v is not None for v in row):
                continue

            item_raw = row[item_idx] if item_idx < len(row) else None
            po_no_raw = row[po_no_idx] if po_no_idx < len(row) else None
            vendor_raw = row[vendor_idx] if vendor_idx < len(row) else None

            if not item_raw or not po_no_raw or not vendor_raw:
                continue

            item_str   = str(item_raw).strip()
            po_no_str  = norm_po_no(po_no_raw)
            vendor_str = str(vendor_raw).strip()

            # Skip placeholder / empty rows
            if not item_str or not po_no_str or not vendor_str:
                continue
            if any(skip in item_str.upper() for skip in ["NOT REQUIRED","NOT ISSUED","NA","N/A"]):
                continue
            if not po_no_str or po_no_str.endswith("/ ") or po_no_str == "PO / NBF /":
                continue
            if len(po_no_str) < 10:
                continue

            # Dates
            po_date_val  = row[po_date_i] if po_date_i < len(row) else None
            indent_date  = row[date_idx] if date_idx < len(row) else None
            po_date_use  = po_date_val if po_date_val else indent_date

            fy = fy_from_date(po_date_use) or sheet_fy

            # Numerics
            qty   = row[qty_idx]   if qty_idx < len(row) else None
            price = row[price_idx] if price_idx < len(row) else None
            gst_p = row[gst_p_idx] if gst_p_idx < len(row) else None
            gst_c = row[gst_c_idx] if gst_c_idx < len(row) else None
            net   = row[net_idx]   if net_idx < len(row) else None
            uom   = row[uom_idx]   if uom_idx < len(row) and uom_idx >= 0 else None

            gst_pct = norm_gst(gst_p)

            # Determine material type from item name heuristics
            item_upper = item_str.upper()
            if any(w in item_upper for w in ["VAC","VACCINE","KILLED","LIVE"]):
                mat_type = "Vaccine"
            elif any(w in item_upper for w in ["MEDICINE","MG BAC","ANTIBIOTIC","ENRO","TYLOSIN"]):
                mat_type = "Medicine"
            elif any(w in item_upper for w in ["VITAMIN","PREMIX","MINERAL","ENZYME","PROBIOTIC","ELECTRO",
                                                 "TRACEMIN","MASTERSORB","KEMZYME","AQUAMAX","SAFEGARD",
                                                 "VOLAMEL","SPERMED","VALOSIN","KOLIN","BIOCHOLINE",
                                                 "FLYVIN","ALKAKARB","DLM","MCP","HRA","BIOTIN"]):
                mat_type = "Feed Medicine"
            else:
                mat_type = "Feed Raw Material"

            # Vendor tracking for parties table
            vkey = vendor_str.lower().strip()
            if vkey not in vendors:
                vendors[vkey] = {"name": vendor_str, "category": mat_type}

            po_records.append({
                "po_no":           po_no_str,
                "po_date":         sdate(po_date_use),
                "fiscal_year":     fy,
                "vendor_name":     vendor_str,
                "item_name":       item_str,
                "material_type":   mat_type,
                "quantity":        snum(qty),
                "unit":            str(uom).strip() if uom else "NULL",
                "rate":            snum(price),
                "gst_pct":         str(gst_pct) if gst_pct else "NULL",
                "total_amount":    snum(net),
                "material_status": "'Pending'",
            })

    return po_records, list(vendors.values())

# ── generate SQL ─────────────────────────────────────────────────────────────

def generate_sql(po_records, vendors):
    lines = []
    lines.append("-- Auto-generated PO import from NBF_ORDER_At_A_Glance.xlsx")
    lines.append("BEGIN;")

    # 1. Parties upsert (vendor names + category)
    lines.append("\n-- Vendors / Parties")
    for v in vendors:
        cat = v["category"]
        lines.append(
            f"INSERT INTO public.parties (name, type, category, is_active) "
            f"VALUES ({sq(v['name'])}, 'supplier', {sq(cat)}, true) "
            f"ON CONFLICT (lower(trim(name)), type) DO UPDATE SET "
            f"category = EXCLUDED.category WHERE public.parties.category IS NULL;"
        )

    # 2. purchase_orders upsert in batches
    lines.append("\n-- Purchase Orders")
    cols = ["po_no","po_date","fiscal_year","vendor_name","item_name","material_type",
            "quantity","unit","rate","gst_pct","total_amount","material_status"]

    batch_size = 100
    for i in range(0, len(po_records), batch_size):
        batch = po_records[i:i+batch_size]
        vals_list = []
        for r in batch:
            unit_val = f"'{r['unit']}'" if r['unit'] != 'NULL' else 'NULL'
            vals_list.append(
                f"({sq(r['po_no'])},{r['po_date']},{sq(r['fiscal_year'])},"
                f"{sq(r['vendor_name'])},{sq(r['item_name'])},{sq(r['material_type'])},"
                f"{r['quantity']},{unit_val},{r['rate']},{r['gst_pct']},"
                f"{r['total_amount']},{r['material_status']})"
            )
        vals_str = ",\n  ".join(vals_list)
        lines.append(
            f"INSERT INTO public.purchase_orders ({','.join(cols)}) VALUES\n  {vals_str}\n"
            f"ON CONFLICT (po_no, item_name) DO UPDATE SET\n"
            f"  po_date=EXCLUDED.po_date, fiscal_year=EXCLUDED.fiscal_year,\n"
            f"  vendor_name=EXCLUDED.vendor_name, material_type=EXCLUDED.material_type,\n"
            f"  quantity=EXCLUDED.quantity, unit=EXCLUDED.unit, rate=EXCLUDED.rate,\n"
            f"  gst_pct=EXCLUDED.gst_pct, total_amount=EXCLUDED.total_amount;"
        )

    lines.append("\nCOMMIT;")
    return "\n".join(lines)

# ── run SQL via Supabase Management API ──────────────────────────────────────

def run_sql(sql):
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        json.dump({"query": sql}, f)
        tmp = f.name
    r = subprocess.run([
        "curl", "-s", "-X", "POST",
        f"https://api.supabase.com/v1/projects/{REF}/database/query",
        "-H", f"Authorization: Bearer {MGMT}",
        "-H", "Content-Type: application/json",
        "--data-binary", f"@{tmp}"
    ], capture_output=True, text=True, timeout=120)
    os.unlink(tmp)
    if r.returncode != 0:
        print("CURL ERROR:", r.stderr)
        sys.exit(1)
    try:
        resp = json.loads(r.stdout)
        if isinstance(resp, dict) and resp.get("error"):
            print("SQL ERROR:", resp["error"])
            sys.exit(1)
        if isinstance(resp, dict) and resp.get("message"):
            print("API ERROR:", resp["message"])
            sys.exit(1)
    except json.JSONDecodeError:
        if "error" in r.stdout.lower():
            print("Response:", r.stdout[:500])
            sys.exit(1)
    return resp

# ── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Reading {XLSX.name} ...")
    po_records, vendors = parse_excel(XLSX)
    print(f"  Parsed {len(po_records)} PO line items across all FY sheets")
    print(f"  Found {len(vendors)} unique vendors")

    if not po_records:
        print("No records found — check Excel format")
        sys.exit(1)

    sql = generate_sql(po_records, vendors)

    # Save SQL for inspection
    sql_path = Path(__file__).parent.parent / "data" / "nbf_po_import.sql"
    sql_path.write_text(sql)
    print(f"  SQL saved to {sql_path.name} ({len(sql)//1024}KB)")

    # Split into chunks (Supabase Management API has ~1MB query limit)
    # Run parties first, then POs in smaller batches
    print("Upserting vendors into parties table ...")
    vendor_lines = [l for l in sql.split("\n") if l.startswith("INSERT INTO public.parties")]
    if vendor_lines:
        vendor_sql = "BEGIN;\n" + "\n".join(vendor_lines) + "\nCOMMIT;"
        run_sql(vendor_sql)
        print(f"  {len(vendor_lines)} vendors upserted")

    print("Importing purchase orders ...")
    po_lines = sql.split("INSERT INTO public.purchase_orders")
    po_chunks = [c for c in po_lines if c.strip() and "ON CONFLICT" in c]
    total = 0
    for idx, chunk in enumerate(po_chunks):
        batch_sql = f"BEGIN;\nINSERT INTO public.purchase_orders{chunk}\nCOMMIT;"
        run_sql(batch_sql)
        total += 100  # approx per chunk
        print(f"  Batch {idx+1}/{len(po_chunks)} done")

    print(f"\nDone. {len(po_records)} PO records imported, {len(vendors)} vendors added to parties.")
